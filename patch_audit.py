import paramiko
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
import os

load_dotenv()

ARQUIVO_SERVIDORES = "servidores.txt"
ARQUIVO_SAIDA = "relatorio_updates.xlsx"

MAX_THREADS = 10

credentials = []

for i in range(1, 20):
    user = os.getenv(f"SSH_USER_{i}")
    password = os.getenv(f"SSH_PASS_{i}")

    if user and password:
        credentials.append((user, password))


def carregar_servidores():
    with open(ARQUIVO_SERVIDORES, "r", encoding="utf-8") as f:
        return [
            linha.strip()
            for linha in f
            if linha.strip() and not linha.startswith("#")
        ]


def conectar_ssh(servidor):
    ultimo_erro = None

    for usuario, senha in credentials:
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

            ssh.connect(
                hostname=servidor,
                username=usuario,
                password=senha,
                timeout=10,
                banner_timeout=10,
                auth_timeout=10,
                look_for_keys=False,
                allow_agent=False,
            )

            return ssh, usuario

        except Exception as erro:
            ultimo_erro = erro

    raise Exception(f"Falha ao conectar. Último erro: {ultimo_erro}")


def executar_comando(ssh, comando):
    stdin, stdout, stderr = ssh.exec_command(comando, timeout=120)

    saida = stdout.read().decode("utf-8", errors="ignore").strip()
    erro = stderr.read().decode("utf-8", errors="ignore").strip()

    return saida, erro


def comando_existe(ssh, comando):
    saida, _ = executar_comando(
        ssh,
        f"command -v {comando} >/dev/null 2>&1 && echo OK || echo NOK"
    )

    return saida.strip() == "OK"


def detectar_so(ssh):
    saida, _ = executar_comando(
        ssh,
        "cat /etc/os-release 2>/dev/null || true"
    )

    texto = saida.lower()

    if "ubuntu" in texto or "debian" in texto:
        return "ubuntu"

    if "suse" in texto or "sles" in texto or "opensuse" in texto:
        return "suse"

    return "desconhecido"


def obter_ip_servidor(ssh, servidor):
    saida, _ = executar_comando(
        ssh,
        "hostname -I | awk '{print $1}'"
    )

    return saida.strip() if saida else servidor


def obter_kernel_atual(ssh):
    saida, _ = executar_comando(ssh, "uname -r")
    return saida.strip() if saida else ""


def verificar_updates_ubuntu(ssh):
    if not comando_existe(ssh, "apt"):
        return []

    executar_comando(
        ssh,
        "sudo -n apt update >/dev/null 2>&1 || true"
    )

    comando = (
        r"apt list --upgradable 2>/dev/null "
        r"| tail -n +2 "
        r"| awk -F/ '{print $1}'"
    )

    saida, _ = executar_comando(ssh, comando)

    return [
        linha.strip()
        for linha in saida.splitlines()
        if linha.strip()
    ]


def verificar_updates_suse(ssh):
    if not comando_existe(ssh, "zypper"):
        return []

    executar_comando(
        ssh,
        "sudo -n zypper --non-interactive refresh >/dev/null 2>&1 || true"
    )

    comando = r"""
zypper --non-interactive --no-refresh list-updates 2>/dev/null \
| grep -E '^[[:space:]]*[vS][[:space:]]*\|' \
| awk -F'|' '{gsub(/^[ \t]+|[ \t]+$/, "", $3); print $3}'
"""

    saida, _ = executar_comando(ssh, comando)

    return [
        linha.strip()
        for linha in saida.splitlines()
        if linha.strip()
    ]


def verificar_updates_suse_com_versao(ssh):
    if not comando_existe(ssh, "zypper"):
        return []

    comando = r"""
zypper --non-interactive --no-refresh list-updates 2>/dev/null \
| grep -E '^[[:space:]]*[vS][[:space:]]*\|' \
| awk -F'|' '{
    status=$1;
    repo=$2;
    nome=$3;
    versao_atual=$4;
    versao_nova=$5;

    gsub(/^[ \t]+|[ \t]+$/, "", nome);
    gsub(/^[ \t]+|[ \t]+$/, "", versao_atual);
    gsub(/^[ \t]+|[ \t]+$/, "", versao_nova);

    print nome " -> " versao_nova
}'
"""

    saida, _ = executar_comando(ssh, comando)

    return [
        linha.strip()
        for linha in saida.splitlines()
        if linha.strip()
    ]


def identificar_updates_kernel(pacotes):
    termos_kernel = [
        "kernel",
        "linux-image",
        "linux-headers",
        "linux-modules",
        "linux-modules-extra",
        "linux-generic",
        "kernel-default",
        "kernel-firmware",
    ]

    return [
        pacote
        for pacote in pacotes
        if any(termo in pacote.lower() for termo in termos_kernel)
    ]


def obter_nova_versao_kernel_ubuntu(ssh):
    comando = r"""
apt list --upgradable 2>/dev/null \
| tail -n +2 \
| grep -E 'linux-image|linux-headers|linux-modules|linux-modules-extra|linux-generic' \
| awk '{print $1 " -> " $2}'
"""

    saida, _ = executar_comando(ssh, comando)

    return [
        linha.strip()
        for linha in saida.splitlines()
        if linha.strip()
    ]


def obter_nova_versao_kernel_suse(ssh):
    atualizacoes = verificar_updates_suse_com_versao(ssh)

    return identificar_updates_kernel(atualizacoes)


def coletar_pacotes_por_so(ssh, servidor, so):
    if so == "ubuntu":
        return verificar_updates_ubuntu(ssh), "ubuntu"

    if so == "suse":
        return verificar_updates_suse(ssh), "suse"

    print(f"Aviso: SO não reconhecido em {servidor}. Tentando apt/zypper...")

    pacotes_apt = verificar_updates_ubuntu(ssh)

    if pacotes_apt:
        return pacotes_apt, "ubuntu (detectado por apt)"

    pacotes_zypper = verificar_updates_suse(ssh)

    if pacotes_zypper:
        return pacotes_zypper, "suse (detectado por zypper)"

    if comando_existe(ssh, "apt"):
        return [], "ubuntu (sem updates)"

    if comando_existe(ssh, "zypper"):
        return [], "suse (sem updates)"

    raise Exception("Sistema operacional não reconhecido")


def obter_nova_versao_kernel_por_so(ssh, so_final):
    so_lower = so_final.lower()

    if "ubuntu" in so_lower:
        return obter_nova_versao_kernel_ubuntu(ssh)

    if "suse" in so_lower:
        return obter_nova_versao_kernel_suse(ssh)

    return []


def verificar_servidor(servidor):
    print(f"Iniciando: {servidor}")

    registro = {
        "servidor": servidor,
        "ip": servidor,
        "so": "",
        "usuario": "",
        "quantidade": 0,
        "pacotes": "",
        "atualizacao_kernel": "Não",
        "kernel_atual": "",
        "nova_versao_kernel": "",
        "kernel_qual": "",
        "status": "",
        "data": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    ssh = None

    try:
        ssh, usuario_usado = conectar_ssh(servidor)

        registro["usuario"] = usuario_usado
        registro["ip"] = obter_ip_servidor(ssh, servidor)
        registro["kernel_atual"] = obter_kernel_atual(ssh)

        so_detectado = detectar_so(ssh)

        pacotes, so_final = coletar_pacotes_por_so(
            ssh,
            servidor,
            so_detectado
        )

        kernels = identificar_updates_kernel(pacotes)
        nova_versao_kernel = obter_nova_versao_kernel_por_so(ssh, so_final)

        registro["so"] = so_final
        registro["quantidade"] = len(pacotes)
        registro["pacotes"] = "\n".join(pacotes)
        registro["atualizacao_kernel"] = "Sim" if kernels else "Não"
        registro["kernel_qual"] = "\n".join(kernels)
        registro["nova_versao_kernel"] = "\n".join(nova_versao_kernel)
        registro["status"] = "OK"

    except Exception as erro:
        registro["status"] = f"Erro: {erro}"

    finally:
        if ssh:
            ssh.close()

    print(
        f"Finalizado: {servidor} | "
        f"IP: {registro['ip']} -> {registro['status']}"
    )

    return registro


def criar_planilha(resultados):
    wb = Workbook()

    ws = wb.active
    ws.title = "Atualizações"

    cabecalho = [
        "Servidor",
        "IP",
        "SO",
        "Usuário",
        "Qtd. Pacotes para atualizar",
        "Lista de pacotes",
        "Atualização de Kernel",
        "Kernel Atual",
        "Nova Versão Kernel",
        "Qual Kernel",
        "Data",
    ]

    ws.append(cabecalho)

    for celula in ws[1]:
        celula.font = Font(bold=True)
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for item in resultados:
        pacotes = (
            item["pacotes"].split("\n")
            if item["pacotes"]
            else []
        )

        lista_limitada = pacotes[:20]
        restante = len(pacotes) - 20

        if restante > 0:
            lista_limitada.append(f"... (+{restante} pacotes)")

        ws.append([
            item["servidor"],
            item["ip"],
            item["so"],
            item["usuario"],
            item["quantidade"],
            "\n".join(lista_limitada),
            item["atualizacao_kernel"],
            item["kernel_atual"],
            item["nova_versao_kernel"],
            item["kernel_qual"],
            item["data"],
        ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 45
    ws.column_dimensions["J"].width = 40
    ws.column_dimensions["K"].width = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    ws.row_dimensions[1].height = 20

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    wb.save(ARQUIVO_SAIDA)


def main():
    servidores = carregar_servidores()

    resultados = []
    falhas = []

    print(f"Total de servidores: {len(servidores)}")
    print(f"Executando com {MAX_THREADS} threads...\n")

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        tarefas = {
            executor.submit(verificar_servidor, servidor): servidor
            for servidor in servidores
        }

        for tarefa in as_completed(tarefas):
            resultado = tarefa.result()

            if resultado["status"] == "OK":
                resultados.append(resultado)
            else:
                falhas.append(resultado)

    resultados.sort(
        key=lambda x: (
            x["atualizacao_kernel"] != "Sim",
            -x["quantidade"],
            x["servidor"]
        )
    )

    criar_planilha(resultados)

    print(f"\nRelatório gerado: {ARQUIVO_SAIDA}")

    if falhas:
        print("\nHosts com falha (não incluídos na planilha):")

        for item in falhas:
            print(
                f"- Host: {item['servidor']} | "
                f"IP: {item['ip']} | "
                f"{item['status']}"
            )


if __name__ == "__main__":
    main()
